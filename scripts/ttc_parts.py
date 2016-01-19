import csv
import os
import string
import time
from os import listdir

import xlrd
import openpyxl
import pandas as pd

from west_data import west_import, west_export

global_master_dict = {}

# Customer Parts - Open required workbooks and check against
def ttc_parts(master_files, path):
    # Dictionary of columns
    columns = {
        0: "NEW/MOD",
        1: "Reason for the change",
        2: "TTC Parts No.",
        3: "Old TTC Parts No.",
        4: "Display Parts No.",
        5: "Parts Name (Global Master)",
        6: "Export HS Code",
        7: "Parts Net Weight",
        8: "WEST Office Code (Exp)",
        9: "WEST Section (Exp)",
        10: "Material Tax Class (Exp)",
        11: "Availability Check Group (Exp)",
        12: "Purchase Group (Exp)",
        13: "Colour Code",
        14: "Unit of Measure",
        15: "WEST Office Code (Imp)",
        16: "WEST Section (Imp)",
        17: "Material Tax Class (Imp)",
        18: "Availability Check Group (Imp)",
        19: "Purchase Group (Imp)",
        20: "Remarks 1",
        21: "Remarks 2",
        22: "Remarks 3",
        23: "Remarks 4",
        24: "Remarks 5"
    }

    # Dictionary of required masters for checking
    required = {
        0: "Customer Contract Details Master",
        1: "Parts Master",
        2: "Customer Parts Master",
        3: "Supplier Parts Master",
        4: "Global Parts Master"
    }

    def check_newmod_field(cell_row, cell_col):
        if all(master_files['xl_sheet_main'].cell_value(cell_row, cell_col) != x for x in ('NEW', 'MOD')):
            print ('NEW/MOD check --- Fail')
            update_df(master_files['xl_sheet_main'].cell_value(cell_row, cell_col), columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), 'NA', 'Please check NEW/MOD field for whitespace')

    def check_maximum_length(cell_row, new_mod):
        # Hard code range of columns to check
        working_columns = list(range(2, 25))

        validate_count = 0
        error_fields = []

        # validate if num   ber
        for col_index in working_columns:
            # validate if blank: Fixed (Probably has its own check, unless Y/N)
            if master_files['xl_sheet_main'].cell_value(8, col_index) == '':
                validate_count += 1
                continue

            try:
                # conditional if number
                maximum_length = int(master_files['xl_sheet_main'].cell_value(8, col_index))
                # check if submitted is integer, remove '.0'
                try:
                    if len(str(int(master_files['xl_sheet_main'].cell_value(cell_row, col_index)))) <= maximum_length:
                        validate_count += 1
                        continue
                    else:
                        error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                        continue
                except ValueError:
                    if len(str(master_files['xl_sheet_main'].cell_value(cell_row, col_index))) <= maximum_length:
                        validate_count += 1
                        continue
                    else:
                        error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                        continue
            except ValueError:
                # conditional if date
                if str(master_files['xl_sheet_main'].cell_value(8, col_index)) == 'dd mmm yyyy':
                    try:
                        time.strptime(str(master_files['xl_sheet_main'].cell_value(cell_row, col_index)),"%d %b %Y")
                        validate_count += 1
                        continue
                    except ValueError:
                        if any(col_index == x for x in (9, 14)): # optional column, can be blank
                            if master_files['xl_sheet_main'].cell_value(cell_row, col_index) == '':
                                validate_count += 1
                                continue
                            else:
                                error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                                continue
                        else:
                            error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                            continue
                # conditional if Month
                elif str(master_files['xl_sheet_main'].cell_value(8, col_index)) == 'mmm yyyy':
                    try:
                        time.strptime(str(master_files['xl_sheet_main'].cell_value(cell_row, col_index)),"%b %Y")
                        validate_count += 1
                        continue
                    except ValueError:
                        if col_index == 10: # optional column, can be blank
                            if master_files['xl_sheet_main'].cell_value(cell_row, col_index) == '':
                                validate_count += 1
                                continue
                            else:
                                error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                                continue
                        else:
                            error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                            continue
                # conditional if integer,integer
                # Only checks total length, not decimal places
                elif master_files['xl_sheet_main'].cell_value(8, col_index).find(',') != -1:
                    char_limit = master_files['xl_sheet_main'].cell_value(8, col_index).split(',')
                    try:
                        if len(str(master_files['xl_sheet_main'].cell_value(cell_row, col_index))) <= (int(char_limit[0]) + int(char_limit[1]) + 1):
                            validate_count += 1
                            continue
                        else:
                            error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                            continue
                    except ValueError:
                        error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                        continue
                else:
                    error_fields.append((col_index, master_files['xl_sheet_main'].cell_value(8, col_index)))
                    continue

        if validate_count == len(working_columns) and len(error_fields) == 0:
            print ('Maximum Length / Date format check --- Pass')
            update_df(new_mod, 'ALL', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', 'NA', 'NA', 'All fields are within maximum characters')
        else:
            for tuple in error_fields:
                print ('Maximum Length / Date format check --- Fail')
                update_df(new_mod, columns[tuple[0]], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, tuple[0]), tuple[1], 'Check format and character length of submitted')

    def check_compulsory_fields(cell_row, new_mod):
        # Hard code compulsory fields
        compulsory_fields = [2, 4, 5, 6, 7, 14]

        if all(master_files['xl_sheet_main'].cell_value(cell_row, col_index) != '' for col_index in compulsory_fields):
            print ('Compulsory Fields check --- Pass')
            update_df(new_mod, 'Compulsory Fields', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', 'NA', 'NA', 'All Compulsory Fields filled')
        else:
            for col_index in compulsory_fields:
                if master_files['xl_sheet_main'].cell_value(cell_row, col_index) == '':
                    print ('Compulsory Fields check --- Fail')
                    update_df(new_mod, columns[col_index], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', PRIMARY_KEY_1, 'NA', columns[col_index] + ' is a Compulsory Field')

    # Check for duplicate primary keys
    def ttc_parts_duplicate_key(cell_row, cell_col, new_mod):
        parts_no_list = []
        for row in range(9, master_files['xl_sheet_main'].nrows):
            parts_no_list.append((master_files['xl_sheet_main'].cell_value(row, 0), master_files['xl_sheet_main'].cell_value(row, 2)))

        matches = 0
        for part_no in parts_no_list:
            # Check if part no same modifier
            if master_files['xl_sheet_main'].cell_value(cell_row, 0) == part_no[0]:
                if PRIMARY_KEY_1 == part_no[1]:
                    matches += 1

        if matches == 1:
            print ('Duplicate Key check --- Pass (Primary key is unique in submitted master)')
            update_df(new_mod, 'Primary Key(s)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', PRIMARY_KEY_1, matches, 'Primary key is unique in submitted master')
        elif matches >1:
            print ('Duplicate Key check --- Fail (Primary key is not unique in submitted master)')
            update_df(new_mod, 'Primary Key(s)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', PRIMARY_KEY_1, matches, 'Primary key is not unique in submitted master')

    # TTC Parts No. should not already exist in system, in CCD, SP, CP, no minor ECI, alphanumeric, uppercase
    def ttc_parts_part_no(cell_row, cell_col, new_mod):
        part_no = master_files['xl_sheet_main'].cell_value(cell_row, cell_col)

        comparison_list_1 = []
        for row in range(9, selected['backup_1'].sheet_by_index(0).nrows):
            comparison_list_1.append(selected['backup_1'].sheet_by_index(0).cell_value(row, 2))

        if part_no not in comparison_list_1:
            print ('TTC Part No. check 1 --- Pass')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'Does not exist in system')
        else:
            backup_row_contents = []
            for row in range(9, selected['backup_1'].sheet_by_index(0).nrows):
                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col) == selected['backup_1'].sheet_by_index(0).cell_value(row, 2):
                    for col in range(0, 25): # Hard Code column range
                        backup_row_contents.append(selected['backup_1'].sheet_by_index(0).cell_value(row, col))

            submitted_row_contents = []
            for col in range(0, 25): # Hard Code column range
                submitted_row_contents.append(master_files['xl_sheet_main'].cell_value(cell_row, col))

            discrepancy_reference = []
            for i, cell_value in enumerate(submitted_row_contents):
                if cell_value != backup_row_contents[i] and all(i != x for x in (0, 1, 2)):
                    discrepancy_reference.append(columns[i])

            if len(discrepancy_reference) == 0:
                discrepancy_reference.append('Submitted has no differences from system')

            print ('TTC Part No. check 1 --- Fail (Duplicate in system)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, ', '.join(discrepancy_reference), 'P/N already exists in system')

        comparison_list_2, comparison_list_3, comparison_list_4 = [], [], []
        matches_2, matches_3, matches_4 = 0, 0, 0
        try:
            for row in range(9, additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].nrows):
                comparison_list_2.append(additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3))

            for backup_part_no in comparison_list_2:
                if part_no == backup_part_no:
                    matches_2 += 1

            if matches_2 >= 1:
                print ('TTC Part No. check 2a --- Pass (NEW part found in CCD)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'New part found in Customer Contract Details')
            else:
                print ('TTC Part No. check 2a --- Fail (Part not found in CCD)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'Part not found in Customer Contract Details')
        except KeyError:
            print ('TTC Part No. check 2a --- Fail (No CCD, SP or CP submitted)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'No Customer Contract Details submitted')

        try:
            for row in range(9, additional['TNM_SUPPLIER_PARTS_MASTER'].nrows):
                comparison_list_3.append(additional['TNM_SUPPLIER_PARTS_MASTER'].cell_value(row, 2))

            for backup_part_no in comparison_list_3:
                if part_no == backup_part_no:
                    matches_3 += 1

            if matches_3 >= 1:
                print ('TTC Part No. check 2b --- Pass (NEW part found in SP)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'New part found in Supplier Parts Master')
            else:
                print ('TTC Part No. check 2b --- Fail (Part not found in SP)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'Part not found in Supplier Parts Master')
        except KeyError:
            print ('TTC Part No. check 2b --- Fail (No CCD, SP or CP submitted)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'No Supplier Parts Master submitted')

        try:
            for row in range(9, additional['TNM_CUSTOMER_PARTS_MASTER'].nrows):
                comparison_list_4.append(additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2))

            for backup_part_no in comparison_list_4:
                if part_no == backup_part_no:
                    matches_4 += 1

            if matches_4 >= 1:
                print ('TTC Part No. check 2C --- Pass (NEW part found in CP)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'New part found in Customer Parts Master')
            else:
                print ('TTC Part No. check 2C --- Fail (Part not found in CP)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'Part not found in Customer Parts Master')
        except KeyError:
            print ('TTC Part No. check 2C --- Fail (No CCD, SP or CP submitted)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'No Customer Parts Master submitted')

        if part_no.endswith(string.ascii_uppercase):
            print ('TTC Part No. check 3 --- Prompt (Part could be ECI)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', part_no, 'NA', 'Part could be ECI')
        else:
            print ('TTC Part No. check 3 --- Pass (Part not ECI)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'Part not ECI')

        if part_no.find(string.punctuation) != -1:
            print ('TTC Part No. check 4 --- Fail (Part No. has symbols within)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'Symbols found in P/N')
        else:
            if part_no == part_no.upper() and part_no.isalnum():
                print ('TTC Part No. check 4 --- Pass')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', part_no, 'NA', 'Alphanumeric and uppercase')
            else:
                print ('TTC Part No. check 4 --- Fail (Part No. is not alphanumeric or uppercase)')
                update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', part_no, 'NA', 'Not alphanumeric and uppercase')

    # TTC P/N must be registered in GM, check old p/n, display p/n, part name (GM), unit of measure, company code
    def ttc_parts_global_master(cell_row, cell_col, part_list, gm_dict, new_mod):
        ttc_part_no = str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col))

        if ttc_part_no in part_list:
            print ('Global Master check 1 --- Pass (Registered in Global Master)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', ttc_part_no, 'NA', 'Registered in Global Master')


            # Replace NoneTypes with ''
            global_master_dict[ttc_part_no]['Company'] = [s for s in global_master_dict[ttc_part_no]['Company'] if s != 'None' and s != '']

            if global_master_dict[ttc_part_no]['Old Part Number'] == 'None':
                global_master_dict[ttc_part_no]['Old Part Number'] = ''

            # Check Old Part No.
            if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) == gm_dict[ttc_part_no]['Old Part Number']:
                print ('Global Master check 2 --- Pass (Old Part No. match)')
                update_df(new_mod, columns[3], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), gm_dict[ttc_part_no]['Old Part Number'], 'Old Part No. match')
            else:
                if 'S500' in gm_dict[ttc_part_no]['Company']:
                    print ('Global Master check 2 --- Fail\n(GM: %s Submitted %s)' % (gm_dict[ttc_part_no]['Old Part Number'], master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1)))
                    update_df(new_mod, columns[3], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), gm_dict[ttc_part_no]['Old Part Number'], 'Old Part No. discrepancy, JP-Sourcing Part')
                else:
                    print ('Global Master check 2 --- Fail\n(GM: %s Submitted %s)' % (gm_dict[ttc_part_no]['Old Part Number'], master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1)))
                    update_df(new_mod, columns[3], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), gm_dict[ttc_part_no]['Old Part Number'], 'Old Part No. discrepancy')

            # Check Display Parts No.
            if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2) == gm_dict[ttc_part_no]['Display Part Number']:
                print ('Global Master check 3 --- Pass (Display Part No. match)')
                update_df(new_mod, columns[4], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), gm_dict[ttc_part_no]['Display Part Number'], 'Display Part No. match')
            else:
                if 'S500' in gm_dict[ttc_part_no]['Company']:
                    print ('Global Master check 3 --- Fail\n(GM: %s Submitted %s)' % (gm_dict[ttc_part_no]['Display Part Number'], master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2)))
                    update_df(new_mod, columns[4], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), gm_dict[ttc_part_no]['Display Part Number'], 'Display Part No. discrepancy, JP-Sourcing Part')
                else:
                    print ('Global Master check 3 --- Fail\n(GM: %s Submitted %s)' % (gm_dict[ttc_part_no]['Display Part Number'], master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2)))
                    update_df(new_mod, columns[4], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), gm_dict[ttc_part_no]['Display Part Number'], 'Display Part No. discrepancy')

            # Check Parts Name
            if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3) == gm_dict[ttc_part_no]['Part Name']:
                print ('Global Master check 4 --- Pass (Part Name match)')
                update_df(new_mod, columns[5], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), gm_dict[ttc_part_no]['Part Name'], 'Parts Name match')
            else:
                if 'S500' in gm_dict[ttc_part_no]['Company']:
                    print ('Global Master check 4 --- Fail')
                    update_df(new_mod, columns[5], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), gm_dict[ttc_part_no]['Part Name'], 'Parts Name discrepancy, JP-Sourcing Part')
                else:
                    print ('Global Master check 4 --- Fail')
                    update_df(new_mod, columns[5], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), gm_dict[ttc_part_no]['Part Name'], 'Parts Name discrepancy')

            # Check Unit of Measure
            if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+12) == gm_dict[ttc_part_no]['Base Quantity Unit']:
                print ('Global Master check 5 --- Pass (Unit of Measure match)')
                update_df(new_mod, columns[14], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+12), gm_dict[ttc_part_no]['Base Quantity Unit'], 'Unit of Measure match')
            else:
                print ('Global Master check 5 --- Fail\n(GM: %s Submitted %s)' % (gm_dict[ttc_part_no]['Base Quantity Unit'], master_files['xl_sheet_main'].cell_value(cell_row, cell_col+12)))
                update_df(new_mod, columns[14], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+12), gm_dict[ttc_part_no]['Base Quantity Unit'], 'Unit of Measure discrepancy, cannot be changed')

            # Check Company Code
            submitted_company_code = ['S566', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+6), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+13)]
            submitted_company_code = [s for s in submitted_company_code if s != '']
            if (all(x in gm_dict[ttc_part_no]['Company'] for x in submitted_company_code)):
                print ('Global Master check 6 --- Pass (Company Code match)')
                update_df(new_mod, 'Company Code', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', submitted_company_code, gm_dict[ttc_part_no]['Company'], 'Company Code match')
            else:
                print ('Global Master check 5 --- Fail')
                update_df(new_mod, 'Company Code', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', submitted_company_code, gm_dict[ttc_part_no]['Company'], 'Company Code discrepancy')

        else:
            print ('Global Master check 1 --- Fail (Part not found in Global Master)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', ttc_part_no, 'NA', 'Not found in Global Master')

    # Display Parts Numbers should be same as TTC Parts Number but with hyphen
    def ttc_parts_display_part(cell_row, cell_col, new_mod):
        correct_part_no = master_files['xl_sheet_main'].cell_value(cell_row, cell_col).replace('-','')

        if master_files['xl_sheet_main'].cell_value(cell_row, cell_col-2) == correct_part_no:
            print ('Display Parts No. check --- Pass')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), master_files['xl_sheet_main'].cell_value(cell_row, cell_col-2), 'P/N with hyphen')
        else:
            print ('Display Parts No. check --- Fail (Does not follow recommended guideline)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), master_files['xl_sheet_main'].cell_value(cell_row, cell_col-2), 'Does not follow recommended guideline (P/N with hyphen)')

    # Check if part name is non-english
    def ttc_parts_part_name(cell_row, cell_col, new_mod):
        try:
            master_files['xl_sheet_main'].cell_value(cell_row, cell_col).encode('ascii')
        except UnicodeEncodeError:
            print ('Part Name check --- Fail (Non-english characters within)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), 'NA', 'Non-english characters within')
        else:
            print ('Part Name check --- Pass (No special characters)')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), 'NA', 'No special characters')

    # Must be alphanumeric with period
    # FUTURE: check format
    def ttc_parts_exp_hs_code(cell_row, cell_col, new_mod):
        exp_hs_code = ''.join(str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col)).split('.'))

        if exp_hs_code.isalnum():
            print ('Exp HS Code check --- Pass (%s)' % master_files['xl_sheet_main'].cell_value(cell_row, cell_col))
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), 'NA', 'Alphanumeric with period')
        else:
            print ('Exp HS Code check --- Fail (Must be alphanumeric with period')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), 'NA', 'Must be alphanumeric with period')

    # NW less than GW/SPQ, Cannot be zero.
    def ttc_parts_net_weight(cell_row, cell_col, new_mod):

        net_weight = float(master_files['xl_sheet_main'].cell_value(cell_row, cell_col))

        customer_parts_spq_gw = []
        for row in range(9, selected['backup_2'].sheet_by_index(0).nrows):
            if selected['backup_2'].sheet_by_index(0).cell_value(row, 2) == master_files['xl_sheet_main'].cell_value(cell_row, cell_col-5):
                customer_parts_spq_gw.append((selected['backup_2'].sheet_by_index(0).cell_value(row, 2), selected['backup_2'].sheet_by_index(0).cell_value(row, 3), selected['backup_2'].sheet_by_index(0).cell_value(row, 10), selected['backup_2'].sheet_by_index(0).cell_value(row, 17)))

        try:
            for row in range(9, additional['TNM_CUSTOMER_PARTS_MASTER'].nrows):
                if additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 0) == 'NEW' and additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2) == master_files['xl_sheet_main'].cell_value(cell_row, cell_col-5):
                    customer_parts_spq_gw.append((additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 3), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 11), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 17)))
                else:
                    for tuple in customer_parts_spq_gw:
                        if str(additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2)) == tuple[0] and additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 3) == tuple[1]:
                            customer_parts_spq_gw.remove(tuple)
                            if additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 11) != '':
                                customer_parts_spq_gw.append((additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 3), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 11), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 17)))
                            else:
                                customer_parts_spq_gw.append((additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 2), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 3), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 10), additional['TNM_CUSTOMER_PARTS_MASTER'].cell_value(row, 17)))

        except KeyError:
            pass

        if len(customer_parts_spq_gw) == 1:
            spq = float(customer_parts_spq_gw[0][2])
            gross_weight = float(customer_parts_spq_gw[0][3])
        else:
            all_spq_gross_weight, spq_list, gross_weight_list = [], [], []
            validate_count = 0
            for tuple in customer_parts_spq_gw:
                spq_list.append(tuple[2])
                gross_weight_list.append(tuple[3])
                all_spq_gross_weight.append((tuple[2], tuple[3]))
            if len(list(set(spq_list))) == 1 and len(list(set(gross_weight_list))) == 1:
                spq = list(set(spq_list))[0]
                gross_weight = list(set(gross_weight_list))[0]
            else:
                for spq, gross_weight in all_spq_gross_weight:
                    if int(gross_weight) / int(spq) >= int(net_weight):
                        validate_count += 1

                if validate_count == len(all_spq_gross_weight):
                    print ('Parts Net Weight check --- Pass (NW: %.5f)' % (net_weight))
                    update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', net_weight, list(set(customer_parts_spq_gw)), 'ALL Gross Weight / SPQ >= Parts Net Weight. Please ensure that all Countries are aware of the changes')
                else:
                    print ('Parts Net Weight check --- Fail')
                    update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', net_weight, list(set(customer_parts_spq_gw)), 'Not ALL Gross Weight / SPQ >= Parts Net Weight. Please ensure that all Countries are aware of the changes, and submit Customer Parts Master for those that fail')

                return

        if gross_weight / spq >= net_weight:
            print ('Parts Net Weight check --- Pass (NW: %.5f)' % (net_weight))
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', net_weight, str([spq, gross_weight]), 'Gross Weight / SPQ >= Parts Net Weight')
        else:
            print ('Parts Net Weight check --- Fail (NW: %.5f)' % (net_weight))
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', net_weight, str([spq, gross_weight]), 'Gross Weight / SPQ < Parts Net Weight')

    # Ensure compulsory WEST fields are input for WEST Countries, NEW Customer Contract Details should be submitted for new WEST Export/Import
    def ttc_parts_west_fields(cell_row, cell_col, new_mod):
        part_no_imp_exp = []
        try:
            for row in range(9, additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].nrows):
                part_no_imp_exp.append((additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 4), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 11)))
        except KeyError:
            if new_mod == 'NEW':
                update_df(new_mod, 'WEST Fields', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), 'NA', 'No Customer Contract Detail submitted for NEW Part/MOD Part with NEW WEST Field')
            else:
                update_df(new_mod, 'WEST Fields', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), 'NA', 'No Customer Contract Detail submitted for MOD Part with NEW WEST Field')

        imp_country_list, exp_country_list = [], []
        for part_customer in part_no_imp_exp:
            if master_files['xl_sheet_main'].cell_value(cell_row, cell_col-6) == part_customer[0]: # Sometimes there's multiple imp countries
                imp_country_list.append(part_customer[1])
                exp_country_list.append(part_customer[2])

        for exp_country in exp_country_list:
            # Check if west export fields are required
            if exp_country[:2] in west_export.keys():
                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col) == west_export[exp_country[:2]]['Office Code']:
                    print ('WEST Export check --- Pass (Office Code match)')
                    update_df(new_mod, columns[8], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), west_export[exp_country[:2]]['Office Code'], 'Office Code match')
                else:
                    print ('WEST Export check --- Fail (Incorrect Office Code %s)' % master_files['xl_sheet_main'].cell_value(cell_row, cell_col))
                    update_df(new_mod, columns[8], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), west_export[exp_country[:2]]['Office Code'], 'Office Code discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) in west_export[exp_country[:2]]['Recommended Section']:
                    print ('WEST Export check --- Pass (Recommended Section match)')
                    update_df(new_mod, columns[9], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), west_export[exp_country[:2]]['Recommended Section'], 'Recommended Section match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) in west_export[exp_country[:2]]['Section']:
                    print ('WEST Export check --- Pass (CAUTION, not recommended section)')
                    update_df(new_mod, columns[9], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), west_export[exp_country[:2]]['Section'], 'Not recommended section')
                else:
                    print ('WEST Export check --- Fail (incorrect Section)')
                    update_df(new_mod, columns[9], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), west_export[exp_country[:2]]['Recommended Section'], 'Section discrepancy')

                if str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2)) == str(west_export[exp_country[:2]]['Material Tax Class']) or str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2)) == str(float(west_export[exp_country[:2]]['Material Tax Class'])):
                    print ('WEST Export check --- Pass (Material Tax Class match)')
                    update_df(new_mod, columns[10], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), west_export[exp_country[:2]]['Material Tax Class'], 'Material Tax Class match')
                else:
                    print ('WEST Export check --- Fail (Incorrect Material Tax Class)')
                    update_df(new_mod, columns[10], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), west_export[exp_country[:2]]['Material Tax Class'], 'Material Tax Class discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3) == 'Z3':
                    print ('WEST Export check --- Pass (Recommended Check Group match)')
                    update_df(new_mod, columns[11], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), 'Z3', 'Recommended Check Group match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3) in west_export[exp_country[:2]]['Availability Check Group']:
                    print ('WEST Export check --- Pass (CAUTION, not recommended check group)')
                    update_df(new_mod, columns[11], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), west_export[exp_country[:2]]['Availability Check Group'], 'Not Recommended Check Group match')
                else:
                    print ('WEST Export check --- Fail (Incorrect Availability Check Group)')
                    update_df(new_mod, columns[11], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), west_export[exp_country[:2]]['Availability Check Group'], 'Availability Check Group discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+4) == west_export[exp_country[:2]]['Purchase Group']:
                    print ('WEST Export check --- Pass (Purchase Group match)')
                    update_df(new_mod, columns[12], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+4), west_export[exp_country[:2]]['Purchase Group'], 'Purchase Group match')
                else:
                    print ('WEST Export check --- Fail (Incorrect Purchase Group)')
                    update_df(new_mod, columns[12], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+4), west_export[exp_country[:2]]['Purchase Group'], 'Purchase Group discrepancy')
            else:
                if (all(field == '' for field in [master_files['xl_sheet_main'].cell_value(cell_row, cell_col), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+4)])):
                    print ('WEST Export check --- Pass (All fields blank)')
                    update_df(new_mod, 'WEST Fields (Exp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), exp_country[:2], 'All fields blank')
                else:
                    print ('WEST Export check --- Warning (fields should be blank)')
                    update_df(new_mod, 'WEST Fields (Exp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+1) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+2) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+3), exp_country[:2], 'WEST fields should be blank for this Export Country. If multiple Export Countries, please check if one of them pertains to this WEST field')

        for imp_country in imp_country_list:
            # Check if west import fields are required
            if imp_country[:2] in west_import.keys() and imp_country[:2] != 'TW':
                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) == west_import[imp_country[:2]]['Office Code']:
                    print ('WEST Import check --- Pass (Office Code match)')
                    update_df(new_mod, columns[15], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7), west_import[imp_country[:2]]['Office Code'], 'Office Code match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Office Code %s)' % master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7))
                    update_df(new_mod, columns[15], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7), west_import[imp_country[:2]]['Office Code'], 'Office Code discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) in west_import[imp_country[:2]]['Recommended Section']:
                    print ('WEST Import check --- Pass (Recommended Section match)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Recommended Section match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) in west_import[imp_country[:2]]['Section']:
                    print ('WEST Import check --- Pass (CAUTION, not recommended section)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Not recommended section')
                else:
                    print ('WEST Import check --- Fail (incorrect Section)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Section discrepancy')

                if str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9)) == str(west_import[imp_country[:2]]['Material Tax Class']) or str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9)) == str(float(west_import[imp_country[:2]]['Material Tax Class'])):
                    print ('WEST Import check --- Pass (Material Tax Class match)')
                    update_df(new_mod, columns[17], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9), west_import[imp_country[:2]]['Material Tax Class'], 'Material Tax Class match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Material Tax Class)')
                    update_df(new_mod, columns[17], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9), west_import[imp_country[:2]]['Material Tax Class'], 'Material Tax Class discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10) == 'Z3':
                    print ('WEST Import check --- Pass (Recommended Check Group match)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), 'Z3', 'Recommended Check Group match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10) in west_import[imp_country[:2]]['Availability Check Group']:
                    print ('WEST Import check --- Pass (CAUTION, not recommended check group)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), west_import[imp_country[:2]]['Availability Check Group'], 'Not Recommended Check Group match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Availability Check Group)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), west_import[imp_country[:2]]['Availability Check Group'], 'Availability Check Group discrepancy')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11) == west_import[imp_country[:2]]['Purchase Group']:
                    print ('WEST Import check --- Pass (Purchase Group match)')
                    update_df(new_mod, columns[19], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11), west_import[imp_country[:2]]['Purchase Group'], 'Purchase Group match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Purchase Group)')
                    update_df(new_mod, columns[19], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11), west_import[imp_country[:2]]['Purchase Group'], 'Purchase Group discrepancy')
            elif imp_country[:2] == 'TW':
                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) == west_import[imp_country[:2]]['Office Code']:
                    print ('WEST Import check --- Pass (Office Code match)')
                    update_df(new_mod, columns[15], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7), west_import[imp_country[:2]]['Office Code'], 'Office Code match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Office Code %s)' % master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7))
                    update_df(new_mod, columns[15], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7), west_import[imp_country[:2]]['Office Code'], 'Office Code discrepancy (Optional for TW)')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) in west_import[imp_country[:2]]['Recommended Section']:
                    print ('WEST Import check --- Pass (Recommended Section match)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Recommended Section match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) in west_import[imp_country[:2]]['Section']:
                    print ('WEST Import check --- Pass (CAUTION, not recommended section)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Not recommended section')
                else:
                    print ('WEST Import check --- Fail (incorrect Section)')
                    update_df(new_mod, columns[16], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), west_import[imp_country[:2]]['Recommended Section'], 'Section discrepancy (Optional for TW)')

                if str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9)) == str(west_import[imp_country[:2]]['Material Tax Class']) or str(master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9)) == str(float(west_import[imp_country[:2]]['Material Tax Class'])):
                    print ('WEST Import check --- Pass (Material Tax Class match)')
                    update_df(new_mod, columns[17], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9), west_import[imp_country[:2]]['Material Tax Class'], 'Material Tax Class match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Material Tax Class)')
                    update_df(new_mod, columns[17], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9), west_import[imp_country[:2]]['Material Tax Class'], 'Material Tax Class discrepancy (Optional for TW)')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10) == 'Z3':
                    print ('WEST Import check --- Pass (Recommended Check Group match)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), 'Z3', 'Recommended Check Group match')
                elif master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10) in west_import[imp_country[:2]]['Availability Check Group']:
                    print ('WEST Import check --- Pass (CAUTION, not recommended check group)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), west_import[imp_country[:2]]['Availability Check Group'], 'Not Recommended Check Group match (Optional for TW)')
                else:
                    print ('WEST Import check --- Fail (Incorrect Availability Check Group)')
                    update_df(new_mod, columns[18], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), west_import[imp_country[:2]]['Availability Check Group'], 'Availability Check Group discrepancy (Optional for TW)')

                if master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11) == west_import[imp_country[:2]]['Purchase Group']:
                    print ('WEST Import check --- Pass (Purchase Group match)')
                    update_df(new_mod, columns[19], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11), west_import[imp_country[:2]]['Purchase Group'], 'Purchase Group match')
                else:
                    print ('WEST Import check --- Fail (Incorrect Purchase Group)')
                    update_df(new_mod, columns[19], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11), west_import[imp_country[:2]]['Purchase Group'], 'Purchase Group discrepancy (Optional for TW)')
            else:
                if (all(field == '' for field in [master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), master_files['xl_sheet_main'].cell_value(cell_row, cell_col+11)])):
                    print ('WEST Import check --- Pass (All fields blank)')
                    update_df(new_mod, 'WEST Fields (Imp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), imp_country[:2], 'All fields blank')
                else:
                    print ('WEST Import check --- Warning (fields should be blank)')
                    update_df(new_mod, 'WEST Fields (Imp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col+7) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+8) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+9) + ', ' + master_files['xl_sheet_main'].cell_value(cell_row, cell_col+10), imp_country[:2], 'WEST fields should be blank for this Import Country. If multiple import countries, please check if one of them pertains to this WEST field')

    # Check Common Parts
    def ttc_parts_common_part(cell_row, cell_col, new_mod):
        part_no_imp_exp = []
        for row in range(9, selected['backup_0'].sheet_by_index(0).nrows):
            if PRIMARY_KEY_1 == selected['backup_0'].sheet_by_index(0).cell_value(row, 3):
                part_no_imp_exp.append((str(selected['backup_0'].sheet_by_index(0).cell_value(row, 3)), selected['backup_0'].sheet_by_index(0).cell_value(row, 6), selected['backup_0'].sheet_by_index(0).cell_value(row, 10)))

        try:
            for row in range(9, additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].nrows):
                if additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 0) == 'NEW' and additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3) == PRIMARY_KEY_1:
                    part_no_imp_exp.append((str(additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3)), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 6), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 10)))
                else: # For MOD parts
                    # Remove existing tuple
                    for tuple in part_no_imp_exp:
                        if additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3) == tuple[0]:
                            part_no_imp_exp.remove(tuple)
                            # Append new tuple
                            part_no_imp_exp.append((str(additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 3)), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 6), additional['TNM_IMP_CUSTOMER_CONTRACT_DETAI'].cell_value(row, 10)))
        except KeyError:
            pass

        import_country, export_country = [], []
        for tuple in part_no_imp_exp:
            import_country.append(tuple[1])
            export_country.append(tuple[2])
        import_country = list(set(import_country)) # Remove duplicates
        export_country = list(set(export_country)) # Remove duplicates

        if len(import_country) == 1:
            print ('Common Parts check 1 --- Pass')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), part_no_imp_exp, 'Only 1 Import Country')
        # TODO: CHECK WHETHER CUSTOMER PARTS MASTER SUBMITED FOR ALL PARTS
        else:
            print ('Common Parts check 1 --- Warning')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), import_country, 'Part is used by multiple Imp Countries, ensure that user has checked with them before changing details')

        if len(export_country) == 1:
            print ('Common Parts check 2 --- Pass')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), export_country, 'Only 1 Export Country')
        else:
            print ('Common Parts check 2 --- Fail')
            update_df(new_mod, columns[cell_col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(cell_row, cell_col), export_country, 'Part is used by multiple Exp Countries, ensure that user has checked with them before changing details')

    # Use colour to test columns to be MOD
    def get_mod_columns(cell_row):
        cols_to_check = []
        for col in range(0, master_files['xl_sheet_main'].ncols):
            # ------- GET COLOUR INDEX ------- #
            xf_list = master_files['xl_workbook'].xf_list[master_files['xl_sheet_main'].cell_xf_index(cell_row, col)]
            cell_font = master_files['xl_workbook'].font_list[xf_list.font_index]
            # ------- GET COLOUR INDEX ------- #

            if all(col != x for x in (0, 1)):
                if cell_font.colour_index == 10 and master_files['xl_sheet_main'].cell_value(cell_row, col) != '':
                    cols_to_check.append(col-2)

        return cols_to_check

    # Check if RED fields are different from system
    # Check if BLACK fields are same as system
    def ttc_parts_mod_reference(cell_row):

        # Get concat key
        part_no = str(master_files['xl_sheet_main'].cell_value(cell_row, 2))

        # Extract all backup concat into list
        comparison_list_1 = []
        for row in range(9, selected['backup_1'].sheet_by_index(0).nrows):
            comparison_list_1.append((row, str(selected['backup_1'].sheet_by_index(0).cell_value(row, 2))))

        # Find backup row
        backup_rows = []
        for concat_str in comparison_list_1:
            if part_no == concat_str[1]:
                backup_rows.append(concat_str[0])

        # If cannot find, return False
        if len(backup_rows) == 0:
            print ('MOD Reference check --- Fail')
            update_df('MOD', 'ALL', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', 'NA', 'NA', 'Cannot find Backup Row in system for MOD part')
            return False

        if len(backup_rows) >= 0:
            # Extract contents of backup row
            backup_row_contents = []
            for col in range(0, 25):
                backup_row_contents.append(selected['backup_1'].sheet_by_index(0).cell_value(backup_rows[0], col))

        submitted_contents = []
        validate_count = 0
        for col in range(0, 25): # Hard Code MAX COLUMN
            submitted_contents.append(master_files['xl_sheet_main'].cell_value(cell_row, col))

            # ------- GET COLOUR INDEX ------- #
            xf_list = master_files['xl_workbook'].xf_list[master_files['xl_sheet_main'].cell_xf_index(cell_row, col)]
            cell_font = master_files['xl_workbook'].font_list[xf_list.font_index]
            # ------- GET COLOUR INDEX ------- #

            if all(col != x for x in (0, 1, 8, 9, 10, 11, 12, 15, 16, 17, 18, 19)): # Don't check WEST Fields
                # IF RED
                if cell_font.colour_index == 10:
                    try:
                        if float(master_files['xl_sheet_main'].cell_value(cell_row, col)) != float(backup_row_contents[col]):
                            validate_count += 1
                        else:
                            print ('MOD Reference Check --- Fail (%s RED but not MOD)' % columns[col])
                            update_df('MOD', columns[col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, col), backup_row_contents[col], 'Field is indicated as \'TO MOD (RED)\' but identical to system')
                    except ValueError:
                        if master_files['xl_sheet_main'].cell_value(cell_row, col) != backup_row_contents[col]:
                            validate_count += 1
                        elif all(x == '' for x in (master_files['xl_sheet_main'].cell_value(cell_row, col), backup_row_contents[col])):
                            validate_count += 1
                        else:
                            print ('MOD Reference Check --- Fail (%s RED but not MOD)' % columns[col])
                            update_df('MOD', columns[col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, col), backup_row_contents[col], 'Field is indicated as \'TO MOD (RED)\' but identical to system')
                # IF BLACK
                elif any(cell_font.colour_index == x for x in (8, 32767)):
                    if master_files['xl_sheet_main'].cell_value(cell_row, col) == backup_row_contents[col]:
                        validate_count += 1
                    elif master_files['xl_sheet_main'].cell_value(cell_row, col) == '' and backup_row_contents[col] != '':
                        validate_count += 1
                    else:
                        try:
                            if float(master_files['xl_sheet_main'].cell_value(cell_row, col)) == float(backup_row_contents[col]):
                                validate_count += 1
                            else:
                                print ('MOD Reference Check --- Fail (%s BLACK but MOD)' % columns[col])
                                update_df('MOD', columns[col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, col), backup_row_contents[col], 'Field is indicated as \'UNCHANGED (BLACK)\' but different from system')
                        except ValueError:
                            print ('MOD Reference Check --- Fail (%s BLACK but MOD)' % columns[col])
                            update_df('MOD', columns[col], cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', master_files['xl_sheet_main'].cell_value(cell_row, col), backup_row_contents[col], 'Field is indicated as \'UNCHANGED (BLACK)\' but different from system')

        if validate_count == len(list(range(2, 8)) + list(range(13, 15)) + list(range(20, 25))): # Hard Code MAX COLUMN
            print ('MOD Reference check --- Pass')
            update_df('MOD', 'ALL', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', str(submitted_contents), str(backup_row_contents), 'Fields are correctly coloured to indicate \'TO CHANGE\'')

        return True

    def ttc_parts_west_mod_reference(cell_row):
        # Get concat key
        part_no = str(master_files['xl_sheet_main'].cell_value(cell_row, 2))

        exp_west_fields = str(master_files['xl_sheet_main'].cell_value(cell_row, 8)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 9)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 10)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 11)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 12))

        imp_west_fields = str(master_files['xl_sheet_main'].cell_value(cell_row, 15)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 16)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 17)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 18)) + str(master_files['xl_sheet_main'].cell_value(cell_row, 19))

        # Find WEST Fields in system
        backup_exp_west_fields, backup_imp_west_fields = [], []
        for row in range(9, selected['backup_1'].sheet_by_index(0).nrows):
            if selected['backup_1'].sheet_by_index(0).cell_value(row, 2) == part_no:
                backup_exp_west_fields.append(str(selected['backup_1'].sheet_by_index(0).cell_value(row, 8)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 9)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 10)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 11)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 12)))

                backup_imp_west_fields.append(str(selected['backup_1'].sheet_by_index(0).cell_value(row, 15)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 16)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 17)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 18)) + str(selected['backup_1'].sheet_by_index(0).cell_value(row, 19)))

        # Check WEST Exp
        exp_west_mod_count = 0
        for col in range(8, 13):
            # ------- GET COLOUR INDEX ------- #
            xf_list = master_files['xl_workbook'].xf_list[master_files['xl_sheet_main'].cell_xf_index(cell_row, col)]
            cell_font = master_files['xl_workbook'].font_list[xf_list.font_index]
            # ------- GET COLOUR INDEX ------- #

            if cell_font.colour_index == 10:
                exp_west_mod_count += 1

        # Check WEST Imp
        imp_west_mod_count = 0
        for col in range(15, 20):
            # ------- GET COLOUR INDEX ------- #
            xf_list = master_files['xl_workbook'].xf_list[master_files['xl_sheet_main'].cell_xf_index(cell_row, col)]
            cell_font = master_files['xl_workbook'].font_list[xf_list.font_index]
            # ------- GET COLOUR INDEX ------- #

            if cell_font.colour_index == 10:
                imp_west_mod_count += 1

        if imp_west_mod_count == 5: # If all Imp WEST Fields are being MOD
            if imp_west_fields not in backup_imp_west_fields:
                print ('Imp WEST Field mod reference check --- Pass')
                update_df('MOD', 'WEST Fields (Imp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', imp_west_fields, ', '.join(list(set(backup_imp_west_fields))), 'Imp WEST Fields validated to MOD')
            else:
                print ('Imp WEST Field mod reference check --- Fail')
                update_df('MOD', 'WEST Fields (Imp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', imp_west_fields, ', '.join(list(set(backup_imp_west_fields))), 'Imp WEST Fields already registered in system')

        if exp_west_mod_count == 5: # If all Exp WEST Fields are being MOD
            if exp_west_fields not in backup_exp_west_fields:
                print ('Exp WEST Field mod reference check --- Pass')
                update_df('MOD', 'WEST Fields (Exp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', exp_west_fields, ', '.join(list(set(backup_exp_west_fields))), 'Exp WEST Fields validated to MOD')
            else:
                if imp_west_fields not in backup_imp_west_fields:
                    print ('Exp WEST Field mod reference check --- Pass')
                    update_df('MOD', 'WEST Fields (Imp/Exp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'PASS', exp_west_fields + ', ' + imp_west_fields, ', '.join(list(set(backup_exp_west_fields)) + list(set(backup_imp_west_fields))), 'Imp/Exp WEST Fields validated to MOD')
                else:
                    print ('Exp WEST Field mod reference check --- Fail')
                    update_df('MOD', 'WEST Fields (Imp/Exp)', cell_row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', exp_west_fields + ', ' + imp_west_fields, ', '.join(list(set(backup_exp_west_fields)) + list(set(backup_imp_west_fields))), 'Imp/Exp WEST Field combination already registered in system')

    # Print required masters for checking

    check_dict = {
        'NEW/MOD': [],
        'Field': [],
        'Row': [],
        'Primary Key': [],
        'Primary Key (Alt)': [],
        'Results': [],
        'Submitted': [],
        'Reference': [],
        'Reason': []
    }

    # Update df dictionary row
    def update_df(new_mod, field, row, primary_key, primary_key_alt, results, submitted, reference, reason):
        check_dict['NEW/MOD'].append(new_mod)
        check_dict['Field'].append(field)
        check_dict['Row'].append(row)
        check_dict['Primary Key'].append(primary_key)
        check_dict['Primary Key (Alt)'].append(primary_key_alt)
        check_dict['Results'].append(results)
        check_dict['Submitted'].append(submitted)
        check_dict['Reference'].append(reference)
        check_dict['Reason'].append(reason)

    print ('The following master files are required: ')
    for i, key in enumerate(required):
        print ('%d: %s' % (i, required.get(key)))
    print ('*' * 60)

    if len(listdir(path + '\\2) Backup')) == 0:
        print ('The folder \'2) Backup\' is empty!')
    else:
        # Automatic loading of backup
        backup_files = []
        for filename in listdir(path + '\\2) Backup\\'):
            backup_files.append(filename)

        # Automatic loading of results if checking GM results
        post_gm_update = input('Would you like to reference post-GM update results? ')
        if post_gm_update == 'Y':
            results_files = []
            for filename in listdir(path + '\\4) Result\\'):
                results_files.append(filename)

        selected = {}
        global_master_part_no = []
        global_master_dict = {}
        for file in backup_files:
            if file.endswith('.xls') or file.endswith('.csv') or file.endswith('xlsx'):
                if file.find('MRS_CustomerContractDetail') != -1:
                    selected['backup_0'] = xlrd.open_workbook(path + '\\2) Backup\\' + file)
                    print ('Retrieved file: %s' % file)
                if file.find('MRS_PartMaster') != -1:
                    selected['backup_1'] = xlrd.open_workbook(path + '\\2) Backup\\' + file)
                    print ('Retrieved file: %s' % file)
                if file.find('MRS_CustomerPartsMaster') != -1:
                    selected['backup_2'] = xlrd.open_workbook(path + '\\2) Backup\\' + file)
                    print ('Retrieved file: %s' % file)
                if file.find('MRS_SupplierPartsMaster') != -1:
                    selected['backup_3'] = xlrd.open_workbook(path + '\\2) Backup\\' + file)
                    print ('Retrieved file: %s' % file)
                if post_gm_update != 'Y':
                    if file.find('MaterialNumber.csv') != -1 and len(file) == 18: # Only full Material Number
                        selected['backup_4'] = open(path + '\\2) Backup\\' + file)
                        selected['backup_4'].seek(0)
                        global_master = csv.DictReader(selected['backup_4'])
                        for row in global_master:
                            global_master_part_no.append(row['Part Number'])
                            global_master_dict[row['Part Number']] = {
                                'Old Part Number': row['Old Part Number'],
                                'Part Name': row['Part Name'],
                                'Base Quantity Unit': row['Base Quantity Unit'],
                                'Display Part Number': row['Display Part Number'],
                                'Company': [row['Company1'], row['Company2'], row['Company3'], row['Company4'], row['Company5'], row['Company6'], row['Company7'], row['Company8'], row['Company9'], row['Company10']]
                            }
                        print ('Retrieved file: %s' % file)
                    if file.find('MaterialNumber.xlsx') != -1: # .xlsx curated version
                        selected['backup_4'] = openpyxl.load_workbook(path + '\\2) Backup\\' + file)
                        material_number = selected['backup_4'].get_sheet_by_name(selected['backup_4'].sheetnames[0]) # Get first worksheet in workbook
                        for row in range(1, material_number.max_row+1):
                            global_master_part_no.append(material_number.cell(row=row, column=1).value)
                            global_master_dict[material_number.cell(row=row, column=1).value] = {
                                'Old Part Number': str(material_number.cell(row=row, column=5).value),
                                'Part Name': material_number.cell(row=row, column=6).value,
                                'Base Quantity Unit': material_number.cell(row=row, column=7).value,
                                'Display Part Number': material_number.cell(row=row, column=9).value,
                                'Company': [
                                    str(material_number.cell(row=row, column=11).value),
                                    str(material_number.cell(row=row, column=12).value),
                                    str(material_number.cell(row=row, column=13).value),
                                    str(material_number.cell(row=row, column=14).value),
                                    str(material_number.cell(row=row, column=15).value),
                                    str(material_number.cell(row=row, column=16).value),
                                    str(material_number.cell(row=row, column=17).value),
                                    str(material_number.cell(row=row, column=18).value),
                                    str(material_number.cell(row=row, column=19).value),
                                    str(material_number.cell(row=row, column=20).value)
                                ]
                            }
                        print ('Retrieved file: %s' % file)

        if post_gm_update == 'Y':
            for file in results_files:
                if file.find('MaterialNumber.csv') != -1 and len(file) == 18: # Only full Material Number
                    selected['backup_4'] = open(path + '\\4) Result\\' + file)
                    selected['backup_4'].seek(0)
                    global_master = csv.DictReader(selected['backup_4'])
                    for row in global_master:
                        global_master_part_no.append(row['Part Number'])
                        global_master_dict[row['Part Number']] = {
                            'Old Part Number': row['Old Part Number'],
                            'Part Name': row['Part Name'],
                            'Base Quantity Unit': row['Base Quantity Unit'],
                            'Display Part Number': row['Display Part Number'],
                            'Company': [row['Company1'], row['Company2'], row['Company3'], row['Company4'], row['Company5'], row['Company6'], row['Company7'], row['Company8'], row['Company9'], row['Company10']]
                        }
                    print ('Retrieved file: %s' % file)

                if file.find('MaterialNumber.xlsx') != -1 and len(file) == 19: # .xlsx curated version
                    selected['backup_4'] = openpyxl.load_workbook(path + '\\4) Result\\' + file)
                    material_number = selected['backup_4'].get_sheet_by_name(selected['backup_4'].sheetnames[0]) # Get first worksheet in workbook
                    for row in range(1, material_number.max_row+1):
                        global_master_part_no.append(material_number.cell(row=row, column=1).value)
                        global_master_dict[material_number.cell(row=row, column=1).value] = {
                            'Old Part Number': str(material_number.cell(row=row, column=5).value),
                            'Part Name': material_number.cell(row=row, column=6).value,
                            'Base Quantity Unit': material_number.cell(row=row, column=7).value,
                            'Display Part Number': material_number.cell(row=row, column=9).value,
                            'Company': [
                                str(material_number.cell(row=row, column=11).value),
                                str(material_number.cell(row=row, column=12).value),
                                str(material_number.cell(row=row, column=13).value),
                                str(material_number.cell(row=row, column=14).value),
                                str(material_number.cell(row=row, column=15).value),
                                str(material_number.cell(row=row, column=16).value),
                                str(material_number.cell(row=row, column=17).value),
                                str(material_number.cell(row=row, column=18).value),
                                str(material_number.cell(row=row, column=19).value),
                                str(material_number.cell(row=row, column=20).value)
                            ]
                        }

        if len(selected) == len(required):
            print('Successfully loaded all backup masters')
            print ('-' * 60)
            print ()
        else:
            # Allow partial retrieval for MOD parts, only if no NEW entries
            change_list = []
            for row in range(9, master_files['xl_sheet_main'].nrows):
                change_list.append(master_files['xl_sheet_main'].cell_value(row, 0))

            if (all(fields != 'NEW' for fields in change_list)):
                print ('WARNING: Not all masters found, program might crash if not all masters available')
                input_to_check = input("Please enter 'Y' to continue, any other key to exit: ")
                if input_to_check != 'Y':
                    return
            else:
                print('Not all masters found, exiting application...')
                return

        # List additional submitted masters to load into memory
        additional = {}
        print ('Loading other submitted masters into memory...')
        print ('Retrieved worksheets:')
        for i, sheet in enumerate(master_files['xl_workbook'].sheets()):
            additional[master_files['xl_workbook'].sheet_by_index(i).name] = master_files['xl_workbook'].sheet_by_index(i)
            print ('%s' % master_files['xl_workbook'].sheet_by_index(i).name)
        print ('-' * 60)
        print ()

        # Checkpoints
        for row in range(9, master_files['xl_sheet_main'].nrows):

            # Set constant for primary keys per master check
            PRIMARY_KEY_1 = master_files['xl_sheet_main'].cell_value(row, 2)
            PRIMARY_KEY_2 = 'NA'

            # If blank, skip row
            if master_files['xl_sheet_main'].cell_value(row, 2) == '':
                continue

            # Print Part no. header
            print ('%s: Part No. %s' % (str(master_files['xl_sheet_main'].cell_value(row, 0)), str(master_files['xl_sheet_main'].cell_value(row, 2))))
            print()

            # If float, prompt user to change format to string
            if isinstance(master_files['xl_sheet_main'].cell_value(row, 2), float):
                print ('Please change format of Row %d TTC Part No. to Text' % row)
                print ('-' * 10)
                update_df('NA', columns[2], row, master_files['xl_sheet_main'].cell_value(row, 2), 'NA', 'FAIL', master_files['xl_sheet_main'].cell_value(row, 2), 'NA', 'Formatting of Row ' + str(row + 1) + ' should be text')
                continue

            check_newmod_field(row, 0)

            # Conditional for NEW parts
            if str(master_files['xl_sheet_main'].cell_value(row, 0)).strip(' ') == 'NEW':
                check_maximum_length(row, 'NEW')
                check_compulsory_fields(row, 'NEW')
                ttc_parts_duplicate_key(row, 2, 'NEW')
                ttc_parts_part_no(row, 2, 'NEW')
                ttc_parts_global_master(row, 2, global_master_part_no, global_master_dict, 'NEW')
                ttc_parts_display_part(row, 4, 'NEW')
                ttc_parts_part_name(row, 5, 'NEW')
                ttc_parts_exp_hs_code(row, 6, 'NEW')
                ttc_parts_net_weight(row, 7, 'NEW')
                ttc_parts_west_fields(row, 8, 'NEW')
            # Conditional for MOD parts
            else:
                global_master_check_cycle, west_fields_check_cycle = 0, 0
                cols_to_check = get_mod_columns(row)
                if len(cols_to_check) != 0:
                    print('User wishes to MOD the following columns:')
                    for col in cols_to_check:
                        print('%s: %s' % (columns[col+2], master_files['xl_sheet_main'].cell_value(row, col+2)))
                    print()

                    if ttc_parts_mod_reference(row):
                        ttc_parts_west_mod_reference(row)
                        check_maximum_length(row, 'MOD')
                        check_compulsory_fields(row, 'MOD')
                        ttc_parts_duplicate_key(row, 2, 'MOD')

                        # Column specific checks
                        for col in cols_to_check:
                            # Mod: TTC Parts No.
                            if col+2 == 2:
                                print ('%s cannot be modded' % columns[col+2])
                                update_df('MOD', columns[col+2], row, master_files['xl_sheet_main'].cell_value(row, 2), 'NA', 'FAIL', master_files['xl_sheet_main'].cell_value(row, 2), 'NA', 'Cannot be modded')
                            # Mod: cols that require GM check
                            # 2016-01-19: Added Column 8 and 15 for MOD Company Code as well
                            if (any(col+2 == x for x in (3, 4, 5, 8, 14, 15))):
                                if global_master_check_cycle == 0:
                                    ttc_parts_global_master(row, 2, global_master_part_no, global_master_dict, 'MOD')
                                    global_master_check_cycle += 1
                            # Mod: Display Part No.
                            if col+2 == 4:
                                ttc_parts_display_part(row, 4, 'MOD')
                                ttc_parts_common_part(row, 4, 'MOD')
                            # Mod: Part Name
                            if col+2 == 5:
                                ttc_parts_part_name(row, 5, 'MOD')
                                ttc_parts_common_part(row, 5, 'MOD')
                            # Mod: Exp HS Code
                            if col+2 == 6:
                                ttc_parts_exp_hs_code(row, 6, 'MOD')
                                ttc_parts_common_part(row, 6, 'MOD')
                            # Mod: Parts Net Weight
                            if col+2 == 7:
                                ttc_parts_net_weight(row, 7, 'MOD')
                                ttc_parts_common_part(row, 7, 'MOD')
                            # Mod: WEST fields
                            if (any(col+2 == x for x in (8, 9, 10, 11, 12, 15, 16, 17, 18, 19))):
                                if west_fields_check_cycle == 0:
                                    ttc_parts_west_fields(row, 8, 'MOD')
                                    west_fields_check_cycle += 1
                            # Mod: optional columns
                            if (any(col+2 == x for x in (20, 21, 22, 23, 24))):
                                print ('There is no programmed check for %s' % columns[col+2])
                                update_df('MOD', columns[col+2], row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'WARNING', master_files['xl_sheet_main'].cell_value(row, col+2), 'NA', 'No programmed check')

                else:
                    print ('There is nothing that is being modded.')
                    update_df('MOD', 'NA', row, PRIMARY_KEY_1, PRIMARY_KEY_2, 'FAIL', 'NA', 'NA', 'Nothing is being modded (No field highlighted RED)')
            print ('-' * 10)
        print('*' * 60)

        # Pandas export to excel
        df = pd.DataFrame(check_dict)
        return df
